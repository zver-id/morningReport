import openpyxl as xls
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import Reference, LineChart, AreaChart, BarChart
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border
from openpyxl.worksheet.dimensions import ColumnDimension
from  openpyxl.chart.label import DataLabelList
import pandas as pd
import numpy as np

anchor_cell = "A1"
width_of_chars= 45
height_of_chars = 14
count_of_days_on_chart = 70
first_column_in_book = dict()

def anchor_cell_increment():
    global anchor_cell
    return_cell = anchor_cell
    anchor_cell = int(anchor_cell[1:]) + 32
    anchor_cell = f"A{anchor_cell}"
    return return_cell

def get_dict_of_cell_number(filename):
    statistics = pd.read_excel(filename, sheet_name='tables')
    dict_of_cell_number = dict()
    index = 2   #первая строка - заголовок
    for value in statistics.iloc[:,0]:
        if value is np.NaN:
            index += 1
        else:
            dict_of_cell_number[value] = index
            index += 1
    return dict_of_cell_number



def remove_excel_sheet(name, filename):
    book = xls.load_workbook(filename)
    if name in book.sheetnames:
        data_sheet = book[name]
        book.remove(data_sheet)
        book.save(filename)
        book.close()

def create_excel_sheet(name, filename):
    book = xls.load_workbook(filename)
    if name not in book.sheetnames:
        book.create_sheet(name)
        book.save(filename)
        book.close()

def create_area_chart(book, line_names):
    sheet = book["Графики"]
    last_column = book["tables"].max_column - 1
    start_column = last_column - count_of_days_on_chart

    area_chart = AreaChart(grouping='percentStacked')
    area_chart.title = "Количество не закрытых инцидентов с разбивкой по времени в работе"
    area_chart.anchor = anchor_cell_increment()
    area_chart.width = width_of_chars
    area_chart.height = height_of_chars


    area_chart.dataLabels = DataLabelList()
    area_chart.dataLabels.showVal = True

    for line in line_names:
        row = first_column_in_book[line]
        series = Reference(book["tables"], min_col=start_column, max_col= last_column, min_row=row, max_row=row)
        area_chart.add_data(series, from_rows=True)

    area_chart.series[0].graphicalProperties.solidFill = '98FB98'  # Первая зона зеленая
    area_chart.series[1].graphicalProperties.solidFill = 'EEE8AA'  # Вторая - песчаная
    area_chart.series[2].graphicalProperties.solidFill = 'FFFF00'  # Третья - желтая
    area_chart.series[3].graphicalProperties.solidFill = 'DC143C'  # Четвертая - красная

    category = Reference(book["tables"], min_col=start_column, max_col= last_column, min_row=1, max_row=1)
    area_chart.set_categories(category)
    set_name_for_series_2(area_chart, line_names)

    area_chart.y_axis.scaling.min = 0.6  # нижняя граница графика - 60%
    sheet.add_chart(area_chart)

def create_bar_chart(book, title, line_names):
    sheet = book["Графики"]
    last_column = book["tables"].max_column - 1
    start_column = last_column - count_of_days_on_chart

    chart = BarChart()
    chart.title = title
    chart.anchor = anchor_cell_increment()
    chart.width = width_of_chars
    chart.height = height_of_chars
    chart.type="col"

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    for line in line_names:
        row = first_column_in_book[line]
        series = Reference(book["tables"], min_col=start_column, max_col= last_column, min_row=row, max_row=row)
        chart.add_data(series, from_rows=True)

    set_name_for_series_2(chart, line_names)

    category = Reference(book["tables"], min_col=start_column, max_col= last_column, min_row=1, max_row=1)
    chart.set_categories(category)

    sheet.add_chart(chart)

def style_text_sheet_list(filename, name):
    book = xls.load_workbook(filename)
    sheet = book[name]
    if name == "tables":
        sheet.delete_cols(2)
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 80
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 5
    sheet.column_dimensions["F"].width = 15
    sheet.column_dimensions["G"].width = 80
    # sheet.border = Border(left=side, right=side, bottom=side, top=side)
    book.save(filename)

def style_text_sheet_data_table(filename, name):
    book = xls.load_workbook(filename)
    sheet = book[name]
    if name == "tables":
        sheet.delete_cols(2)
    sheet.column_dimensions["A"].width = 50

    book.save(filename)

def insert_old_tickets(filename, old_tikets):
    book = xls.load_workbook(filename)
    if "Старые обращения" in book.sheetnames:
        remove_excel_sheet("Старые обращения", filename)
    create_excel_sheet("Старые обращения", filename)
    book = xls.load_workbook(filename)
    sheet = book["Старые обращения"]
    for row in old_tikets:
        sheet.append(row)
    sheet.insert_rows(0)
    sheet["A1"].value = "Дата поступления"
    sheet.column_dimensions["A"].width = 20

    sheet["B1"].value = "Номер"
    sheet.column_dimensions["B"].width = 20

    sheet["C1"].value = "Описание"
    sheet.column_dimensions["C"].width = 100

    sheet["D1"].value = "Ответственный"
    sheet.column_dimensions["D"].width = 25

    sheet["E1"].value = "Состояние"
    sheet.column_dimensions["E"].width = 25

    book.save(filename)
    book.close()

def create_line_chart(book, title, line_names):
    sheet = book["Графики"]
    last_column = book["tables"].max_column - 1
    start_column = last_column - count_of_days_on_chart

    chart = LineChart()
    chart.title = title
    chart.anchor = anchor_cell_increment()
    chart.width = width_of_chars
    chart.height = height_of_chars

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    for line in line_names:
        row = first_column_in_book[line]
        series = Reference(book["tables"], min_col=start_column, max_col=last_column, min_row=row, max_row=row)
        chart.add_data(series, from_rows=True)

    category = Reference(book["tables"], min_col=start_column, max_col= last_column, min_row=1, max_row=1)
    chart.set_categories(category)

    set_name_for_series_2(chart, line_names)

    sheet.add_chart(chart)

def set_name_for_series(chart, min_row, max_row):
    if min_row == max_row:
        row_series = chart.series[0]
        title = f"tables!A{min_row}"
        title = SeriesLabel(strRef=StrRef(title))
        row_series.title = title
    else:
        series = 0
        for row in range(min_row, max_row+1):
            row_series = chart.series[series]
            title = f"tables!A{row}"
            title = SeriesLabel(strRef=StrRef(title))
            row_series.title = title
            series +=1

def set_name_for_series_2(chart, line_names):
    if len(line_names) == 1:
        row_series = chart.series[0]
        row = first_column_in_book[line_names[0]]
        title = f"tables!A{row}"
        title = SeriesLabel(strRef=StrRef(title))
        row_series.title = title
    else:
        series = 0
        for line in line_names:
            row_series = chart.series[series]
            row = first_column_in_book[line]
            title = f"tables!A{row}"
            title = SeriesLabel(strRef=StrRef(title))
            row_series.title = title
            series +=1

def write_chars_to_file(filename):
    global first_column_in_book
    first_column_in_book = get_dict_of_cell_number(filename)
    book = xls.load_workbook(filename)
    # create_chart_all_tickets(book)
    chart_options = ['Март 2023', 'Апрель 2023', 'Май 2023', 'Июнь 2023', 'Всего в работе']
    create_line_chart(book, "Количество обращений в состоянии 'В работе'", chart_options)

    print("один записал")
    #create_chart_older_two_weeks(book)
    chart_options = ["Старше 2 недель", "Старше 3 недель"]
    create_line_chart(book, "Snowball", chart_options)
    print("еще один")

    #create_chart_tail(book)
    chart_options = ["Хвост"]
    create_line_chart(book, "Общее количество не закрытых обращений", chart_options)
    print("и еще один записал")

    #create_chart_tail_older_4_weeks(book)
    chart_options = ["Старше 4 недель"]
    create_line_chart(book, "Количество не закрытых обращений старше 4 недель", chart_options)

    chart_options = ["0-8", "8-16", "16-24", ">24"]
    create_area_chart(book, chart_options)

    chart_options = ['Поступившие', 'Проработанные']
    create_bar_chart(book, "Количество поступивших негативных оценок", chart_options)
    print("еще чуть чуть")
    chart_options = ['Кол-во баллов за месяц']
    create_bar_chart(book, "Баллы внешних сообщений", chart_options)
    chart_options = ['Инциденты', 'Консультации', 'Запросы', 'Проблемы']
    create_line_chart(book, "Поступило обращений по типам", chart_options)
    print("Еще крапалиночку")
    chart_options = ['Поступило всего']
    create_line_chart(book, "Поступило обращений всего", chart_options)
    chart_options = ['Запросы', 'Затрачено в часах']
    create_line_chart(book, "Затрачено на запросы", chart_options)
    book.save(filename)
    global anchor_cell
    anchor_cell = "A1"

if __name__ == "__main__":
    write_chars_to_file('data.xlsx')
